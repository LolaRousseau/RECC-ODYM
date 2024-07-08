# -*- coding: utf-8 -*-
"""
Created on Jan 5th, 2020, as copy of RECC_ScenarioControl_V2_2.py

@author: spauliuk
"""

"""

File RECC_ScenarioControl.py

Script that modifies the RECC config file to run a list of scenarios and executes RECC main script for each scenario config.

"""

# Import required libraries:
import os
import openpyxl

import RECC_Paths # Import path file
import ODYM_RECC_Main

#ScenarioSetting, sheet name of RECC_ModelConfig_List.xlsx to be selected:
#ScenarioSetting = 'pav_reb_Config_list'
#ScenarioSetting = 'pav_reb_Config_list_all'
#ScenarioSetting = 'Germany_detail_config'
#ScenarioSetting = 'Germany_detail_config_all'
#ScenarioSetting = 'Global_all'
#ScenarioSetting = 'TestRun'
ScenarioSetting = 'Greater_Oslo'

# open scenario sheet
ModelConfigListFile  = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,'RECC_ModelConfig_List.xlsx'))
ModelConfigListSheet = ModelConfigListFile[ScenarioSetting]
SheetName = 'Config_Auto'
#Read control lines and execute main model script
ResultFolders = []
Row = 3
# search for script config list entry
while ModelConfigListSheet.cell(Row+1, 3).value != None:
    RegionalScope = ModelConfigListSheet.cell(Row+1, 3).value
    print(RegionalScope)
    Config = {}
    for m in range(3,12):
        Config[ModelConfigListSheet.cell(3, m+1).value] = ModelConfigListSheet.cell(Row+1, m+1).value
    Row += 1
    # rewrite RECC model config
    mywb = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,'RECC_Config.xlsx'))
    
    sheet = mywb['Cover']
    sheet['D4'] = SheetName
    sheet = mywb[SheetName]
    sheet['D7']   = RegionalScope
    sheet['G24']  = Config['RegionSelect']
    # sheet['G27']  = Config['Products']
    # sheet['G28']  = Config['Sectors']
    # sheet['G29']  = Config['Products']
    # sheet['G33']  = Config['NonresidentialBuildings']
    # sheet['G48']  = Config['Regions32goods']
    # # The indices below need to be updated when new parameters are added to the parameter list
    # sheet['D183'] = Config['Logging_Verbosity']
    # sheet['D184'] = Config['Include_REStrategy_FabYieldImprovement']
    # sheet['D185'] = Config['Include_REStrategy_FabScrapDiversion']
    # sheet['D186'] = Config['Include_REStrategy_EoL_RR_Improvement']
    # sheet['D187'] = Config['ScrapExport']
    # sheet['D188'] = Config['ScrapExportRecyclingCredit']
    # sheet['D189'] = Config['IncludeRecycling']
    # sheet['D190'] = Config['Include_REStrategy_MaterialSubstitution']
    # sheet['D191'] = Config['Include_REStrategy_UsingLessMaterialByDesign']
    # sheet['D192'] = Config['Include_REStrategy_ReUse']
    # sheet['D193'] = Config['Include_REStrategy_LifeTimeExtension']
    # sheet['D194'] = Config['Include_REStrategy_MoreIntenseUse']
    # sheet['D195'] = Config['Include_REStrategy_CarSharing']
    # sheet['D196'] = Config['Include_REStrategy_RideSharing']
    # sheet['D197'] = Config['Include_REStrategy_ModalSplit']
    sheet['D225'] = Config['SectorSelect']
    # sheet['D199'] = Config['Include_Renovation_reb']
    # sheet['D200'] = Config['Include_Renovation_nrb']
    # sheet['D201'] = Config['No_EE_Improvements']
    # sheet['D215'] = Config['PlotResolution']
    sheet['D231'] = Config['pkm_alternative']
    sheet['D232'] = Config['lifetime_buildings']
    sheet['D233'] = Config['pop_scenario']
    sheet['D234'] = Config['Include_DecarbElectricity']
    sheet['D235'] = Config['Include_Materialstechnologyshare']
    sheet['D236'] = Config['Include_ScenarioFApC']
    sheet['D237'] = Config['Include_ScenarioBuildType']
    
    mywb.save(os.path.join(RECC_Paths.data_path,'RECC_Config.xlsx'))

    # run the ODYM-RECC model
    OutputDict = ODYM_RECC_Main.main()
    ResultFolders.append(OutputDict['Name_Scenario'])

# Export ResultFolders:
book = openpyxl.Workbook()
ws1 = book.active
ws1.title = 'ResultFolders'
Fr = 3
for Fname in ResultFolders:
    ws1.cell(row=Fr+1, column=4).value = Fname 
    Fr +=1
book.save(os.path.join(RECC_Paths.results_path,'ResultFolders.xlsx'))   
#
#
